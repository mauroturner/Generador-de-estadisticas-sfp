import numpy as np
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import pandas as pd
import os 

"""
Estadísticas por 
- Ministerio
- Localidad
- Puesto actual
- Género
"""

def main():
    excel_input = 'Archivos\\Entrada'
    excel_output = 'Archivos\\Salida'
    for archivo in os.listdir(excel_input):
        f = os.path.join(excel_input, archivo)
        excel_output = os.path.join(excel_output, archivo)
        if os.path.isfile(f):
            # Obtenemos los datos y ordenamos
            df = pd.read_excel(f, header=0).sort_values('MINISTERIO / ENTE / ORGANISMO')

            # Agrupamos por ministerio
            estadisticas = df.groupby('MINISTERIO / ENTE / ORGANISMO')

            # Preparamos la hoja de cálculo
            wb = openpyxl.Workbook()
            hoja = wb.active

            # Añadimos encabezados y estilos
            hoja.merge_cells('A1:G1')
            hoja['A1'].value = archivo.split('.')[0]
            hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['A1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            hoja['A1'].font = Font(name='Arial', size=8)
            hoja['A3'].value = 'MINISTERIO / ENTE / ORGANISMO'
            hoja['B3'].value = 'LOCALIDAD'
            hoja.merge_cells('C3:D3')
            hoja['C3'].value = 'FUNCIÓN'
            hoja.merge_cells('E3:F3')
            hoja['E3'].value = 'GÉNERO'
            hoja['G3'].value = 'TOTAL'
            hoja.column_dimensions['A'].width = 25
            hoja.column_dimensions['B'].width = 25
            hoja.column_dimensions['C'].width = 25
            hoja['A3'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['A3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            hoja['A3'].font = Font(name='Arial', size=8)
            hoja['B3'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['B3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            hoja['B3'].font = Font(name='Arial', size=8)
            hoja['C3'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['C3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            hoja['C3'].font = Font(name='Arial', size=8)
            hoja['E3'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['E3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            hoja['E3'].font = Font(name='Arial', size=8)
            hoja['G3'].alignment = Alignment(horizontal='center', vertical='center')
            hoja['G3'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            hoja['G3'].font = Font(name='Arial', size=8)

            # Generamos las estadísticas
            i = 4
            j = 4
            for ministerio, grupo in estadisticas:
                comienza_en_fila = i
                localidad_comienza_en_fila = i
                total_por_ministerio = 0
                por_localidad = estadisticas.get_group(ministerio).groupby('LOCALIDAD')
                for localidad, grupo in por_localidad:
                    por_puesto_actual = por_localidad.get_group(localidad).groupby('PUESTO ACTUAL').aggregate('value_counts').reset_index(name='TOTAL')
                    por_cargo_actual = por_localidad.get_group(localidad).groupby('PUESTO ACTUAL')
                    hoja['B' + str(i)].value = localidad
                    for index, puesto_actual in por_puesto_actual.iterrows():
                        hoja['C' + str(i)].alignment = Alignment(horizontal='center', vertical='center')
                        hoja['C' + str(i)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        hoja['C' + str(i)].font = Font(name='Arial', size=8)
                        hoja['C' + str(i)].value = puesto_actual['PUESTO ACTUAL']
                        hoja['D' + str(i)].alignment = Alignment(horizontal='center', vertical='center')
                        hoja['D' + str(i)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        hoja['D' + str(i)].font = Font(name='Arial', size=8)
                        hoja['D' + str(i)].value = puesto_actual['TOTAL']
                        total_por_ministerio += puesto_actual['TOTAL']
                        i += 1
                    hoja.merge_cells('B' + str(localidad_comienza_en_fila) + ':' + 'B' + str(i-1))
                    hoja['B' + str(localidad_comienza_en_fila)].alignment = Alignment(horizontal='center', vertical='center')
                    hoja['B' + str(localidad_comienza_en_fila)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    hoja['B' + str(localidad_comienza_en_fila)].font = Font(name='Arial', size=8)
                    hoja.merge_cells('G' + str(comienza_en_fila) + ':' + 'G' + str(i-1))
                    hoja['G' + str(comienza_en_fila)].alignment = Alignment(horizontal='center', vertical='center')
                    hoja['G' + str(comienza_en_fila)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    hoja['G' + str(comienza_en_fila)].font = Font(name='Arial', size=8)
                    hoja['G' + str(comienza_en_fila)].font = Font(name='Arial', size=8)
                    hoja['G' + str(comienza_en_fila)].value = total_por_ministerio
                    localidad_comienza_en_fila = i
                    for cargo, grupo in por_cargo_actual:
                        por_genero = por_cargo_actual.get_group(cargo).groupby('GÉNERO').aggregate('value_counts').reset_index(name='TOTAL')
                        for index, genero in por_genero.iterrows():
                            hoja['E' + str(j)].alignment = Alignment(horizontal='center', vertical='center')
                            hoja['E' + str(j)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            hoja['E' + str(j)].font = Font(name='Arial', size=8)
                            hoja['E' + str(j)].value = genero['GÉNERO']
                            hoja['F' + str(j)].alignment = Alignment(horizontal='center', vertical='center')
                            hoja['F' + str(j)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            hoja['F' + str(j)].font = Font(name='Arial', size=8)
                            hoja['F' + str(j)].value = genero['TOTAL']
                            j += 1
                hoja.merge_cells('A' + str(comienza_en_fila) + ':' + 'A' + str(i-1))
                hoja['A' + str(comienza_en_fila)].alignment = Alignment(horizontal='center', vertical='center')
                hoja['A' + str(comienza_en_fila)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                hoja['A' + str(comienza_en_fila)].font = Font(name='Arial', size=8)
                hoja['A' + str(comienza_en_fila)].value = ministerio
            hoja['F' + str(i)].value = 'TOTAL'
            hoja['G' + str(i)].value = str(len(df.index))
            hoja['F' + str(i)].alignment = Alignment(horizontal='center', vertical='center')
            hoja['F' + str(i)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            hoja['F' + str(i)].font = Font(name='Arial', size=8)
            hoja['G' + str(i)].alignment = Alignment(horizontal='center', vertical='center')
            hoja['G' + str(i)].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            hoja['G' + str(i)].font = Font(name='Arial', size=8)
            wb.save(os.path.join('Archivos\\Salida', archivo))
        # limpiamos la ruta
        f = ''

if __name__ == '__main__':
    main()